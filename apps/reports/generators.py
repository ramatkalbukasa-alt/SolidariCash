"""
SolidariCash — Générateurs de rapports PDF et Excel
"""
import io
from decimal import Decimal
from django.conf import settings
from django.utils import timezone


def generate_cycle_pdf(cycle):
    """Génère un rapport PDF complet pour un cycle."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    from apps.contributions.models import Contribution
    from apps.distributions.models import Distribution
    from apps.rotation.models import RotationOrder

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        rightMargin=1.5*cm, leftMargin=1.5*cm,
        topMargin=2*cm, bottomMargin=2*cm
    )

    # Couleurs
    GREEN = colors.HexColor('#22C55E')
    DARK = colors.HexColor('#0F172A')
    LIGHT_GREY = colors.HexColor('#F1F5F9')
    MID_GREY = colors.HexColor('#64748B')

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Title'], textColor=DARK, fontSize=18, spaceAfter=6)
    subtitle_style = ParagraphStyle('Sub', parent=styles['Normal'], textColor=MID_GREY, fontSize=10, spaceAfter=12)
    section_style = ParagraphStyle('Section', parent=styles['Heading2'], textColor=GREEN, fontSize=13, spaceBefore=16, spaceAfter=8)
    normal_style = styles['Normal']

    story = []

    # En-tête
    story.append(Paragraph("SolidariCash", title_style))
    story.append(Paragraph(f"Rapport du cycle : {cycle.name}", subtitle_style))
    story.append(Paragraph(f"Généré le {timezone.now().strftime('%d/%m/%Y à %H:%M')}", subtitle_style))
    story.append(HRFlowable(width="100%", thickness=2, color=GREEN, spaceAfter=16))

    # Informations cycle
    story.append(Paragraph("Informations du cycle", section_style))
    cycle_data = [
        ["Nom du cycle", cycle.name],
        ["Période", f"{cycle.start_date.strftime('%d/%m/%Y')} — {cycle.end_date.strftime('%d/%m/%Y')}"],
        ["Montant par tête", f"{settings.APP_CURRENCY_SYMBOL}{cycle.contribution_amount}"],
        ["Taux de commission", f"{float(cycle.commission_rate) * 100:.1f}%"],
        ["Statut", cycle.get_status_display()],
        ["Total collecté", f"{settings.APP_CURRENCY_SYMBOL}{cycle.total_collected:.2f}"],
        ["Commission", f"{settings.APP_CURRENCY_SYMBOL}{cycle.commission_amount:.2f}"],
        ["Montant distribuable", f"{settings.APP_CURRENCY_SYMBOL}{cycle.distributable_amount:.2f}"],
    ]
    t = Table(cycle_data, colWidths=[6*cm, 10*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), LIGHT_GREY),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('PADDING', (0, 0), (-1, -1), 8),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
    ]))
    story.append(t)

    # Contributions
    story.append(Paragraph("Contributions", section_style))
    contributions = Contribution.objects.filter(cycle=cycle).select_related('head__member__user').order_by('head__member__user__last_name')
    contrib_data = [["Membre", "Tête", "Montant dû", "Montant payé", "Statut"]]
    for c in contributions:
        contrib_data.append([
            c.head.member.full_name,
            f"Tête #{c.head.head_number}",
            f"{settings.APP_CURRENCY_SYMBOL}{c.amount_due:.2f}",
            f"{settings.APP_CURRENCY_SYMBOL}{c.amount_paid:.2f}",
            c.get_status_display(),
        ])
    ct = Table(contrib_data, colWidths=[5*cm, 3*cm, 3*cm, 3.5*cm, 3*cm])
    ct.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), GREEN),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#CBD5E1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, LIGHT_GREY]),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(ct)

    # Rotation
    story.append(Paragraph("Ordre de rotation", section_style))
    orders = RotationOrder.objects.filter(cycle=cycle).select_related('head__member__user').order_by('position')
    rot_data = [["Pos.", "Membre", "Tête", "Statut", "Urgence"]]
    for o in orders:
        rot_data.append([
            str(o.position),
            o.head.member.full_name,
            f"Tête #{o.head.head_number}",
            o.get_status_display(),
            "Oui" if o.is_emergency else "Non",
        ])
    rt = Table(rot_data, colWidths=[2*cm, 6*cm, 3*cm, 4*cm, 2.5*cm])
    rt.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), DARK),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#CBD5E1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, LIGHT_GREY]),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(rt)

    doc.build(story)
    buffer.seek(0)
    return buffer


def generate_cycle_excel(cycle):
    """Génère un rapport Excel complet pour un cycle."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    from apps.contributions.models import Contribution
    from apps.distributions.models import Distribution
    from apps.rotation.models import RotationOrder

    wb = openpyxl.Workbook()

    # Styles
    green_fill = PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")
    dark_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    light_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    header_font = Font(bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1'),
    )

    def style_header(ws, row, cols):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = green_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

    def style_data_row(ws, row, cols, alternate=False):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            if alternate:
                cell.fill = light_fill
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

    # === Feuille Résumé ===
    ws_summary = wb.active
    ws_summary.title = "Résumé"
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 30

    summary_data = [
        ("SolidariCash — Rapport de cycle", ""),
        ("", ""),
        ("Nom du cycle", cycle.name),
        ("Période", f"{cycle.start_date.strftime('%d/%m/%Y')} - {cycle.end_date.strftime('%d/%m/%Y')}"),
        ("Montant par tête", float(cycle.contribution_amount)),
        ("Taux de commission", f"{float(cycle.commission_rate) * 100:.1f}%"),
        ("Statut", cycle.get_status_display()),
        ("Total collecté", float(cycle.total_collected)),
        ("Commission", float(cycle.commission_amount)),
        ("Montant distribuable", float(cycle.distributable_amount)),
    ]
    for i, (k, v) in enumerate(summary_data, 1):
        ws_summary.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws_summary.cell(row=i, column=2, value=v)

    # === Feuille Contributions ===
    ws_contrib = wb.create_sheet("Contributions")
    headers = ["Membre", "Tête", "Montant dû", "Montant payé", "Reste", "Mode paiement", "Statut", "Date validation"]
    for col, h in enumerate(headers, 1):
        ws_contrib.cell(row=1, column=col, value=h)
    style_header(ws_contrib, 1, len(headers))
    ws_contrib.row_dimensions[1].height = 22

    contribs = Contribution.objects.filter(cycle=cycle).select_related('head__member__user')
    for i, c in enumerate(contribs, 2):
        row_data = [
            c.head.member.full_name,
            f"Tête #{c.head.head_number}",
            float(c.amount_due),
            float(c.amount_paid),
            float(c.remaining_amount),
            c.get_payment_method_display() if c.payment_method else '-',
            c.get_status_display(),
            c.validated_at.strftime('%d/%m/%Y %H:%M') if c.validated_at else '-',
        ]
        for col, val in enumerate(row_data, 1):
            ws_contrib.cell(row=i, column=col, value=val)
        style_data_row(ws_contrib, i, len(headers), alternate=(i % 2 == 0))

    for col in range(1, len(headers) + 1):
        ws_contrib.column_dimensions[get_column_letter(col)].width = 18

    # === Feuille Rotation ===
    ws_rot = wb.create_sheet("Rotation")
    rot_headers = ["Position", "Membre", "Tête", "Statut", "Urgence", "Date programmée"]
    for col, h in enumerate(rot_headers, 1):
        ws_rot.cell(row=1, column=col, value=h)
    style_header(ws_rot, 1, len(rot_headers))

    orders = RotationOrder.objects.filter(cycle=cycle).select_related('head__member__user').order_by('position')
    for i, o in enumerate(orders, 2):
        row_data = [
            o.position, o.head.member.full_name,
            f"Tête #{o.head.head_number}", o.get_status_display(),
            "Oui" if o.is_emergency else "Non",
            o.scheduled_date.strftime('%d/%m/%Y') if o.scheduled_date else '-',
        ]
        for col, val in enumerate(row_data, 1):
            ws_rot.cell(row=i, column=col, value=val)
        style_data_row(ws_rot, i, len(rot_headers), alternate=(i % 2 == 0))

    for col in range(1, len(rot_headers) + 1):
        ws_rot.column_dimensions[get_column_letter(col)].width = 18

    # === Feuille Distributions ===
    ws_dist = wb.create_sheet("Distributions")
    dist_headers = ["Membre", "Tête", "Montant brut", "Commission", "Montant net", "Statut", "Traité le"]
    for col, h in enumerate(dist_headers, 1):
        ws_dist.cell(row=1, column=col, value=h)
    style_header(ws_dist, 1, len(dist_headers))

    dists = Distribution.objects.filter(cycle=cycle).select_related('head__member__user')
    for i, d in enumerate(dists, 2):
        row_data = [
            d.head.member.full_name, f"Tête #{d.head.head_number}",
            float(d.gross_amount), float(d.commission_amount), float(d.net_amount),
            d.get_status_display(),
            d.processed_at.strftime('%d/%m/%Y %H:%M') if d.processed_at else '-',
        ]
        for col, val in enumerate(row_data, 1):
            ws_dist.cell(row=i, column=col, value=val)
        style_data_row(ws_dist, i, len(dist_headers), alternate=(i % 2 == 0))

    for col in range(1, len(dist_headers) + 1):
        ws_dist.column_dimensions[get_column_letter(col)].width = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_member_receipt_pdf(distribution):
    """Génère un reçu PDF pour une distribution."""
    from reportlab.lib.pagesizes import A5
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.enums import TA_CENTER

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A5, rightMargin=1.5*cm, leftMargin=1.5*cm, topMargin=2*cm, bottomMargin=2*cm)

    GREEN = colors.HexColor('#22C55E')
    DARK = colors.HexColor('#0F172A')
    styles = getSampleStyleSheet()
    center_style = ParagraphStyle('Center', parent=styles['Normal'], alignment=TA_CENTER, fontSize=10)
    title_style = ParagraphStyle('Title', parent=styles['Title'], alignment=TA_CENTER, textColor=GREEN, fontSize=16)

    story = []
    story.append(Paragraph("SolidariCash", title_style))
    story.append(Paragraph("REÇU DE DISTRIBUTION", ParagraphStyle('Sub', parent=styles['Normal'], alignment=TA_CENTER, fontSize=11, textColor=DARK)))
    story.append(Spacer(1, 0.5*cm))
    story.append(HRFlowable(width="100%", thickness=2, color=GREEN, spaceAfter=12))

    data = [
        ["Bénéficiaire", distribution.head.member.full_name],
        ["Tête", distribution.head.display_name],
        ["Cycle", distribution.cycle.name],
        ["Montant brut", f"{settings.APP_CURRENCY_SYMBOL}{distribution.gross_amount:.2f}"],
        ["Commission", f"{settings.APP_CURRENCY_SYMBOL}{distribution.commission_amount:.2f}"],
        ["Montant net", f"{settings.APP_CURRENCY_SYMBOL}{distribution.net_amount:.2f}"],
        ["Date", distribution.processed_at.strftime('%d/%m/%Y %H:%M') if distribution.processed_at else '-'],
        ["Statut", distribution.get_status_display()],
    ]
    t = Table(data, colWidths=[5*cm, 7*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F1F5F9')),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('PADDING', (0, 0), (-1, -1), 8),
        ('BACKGROUND', (0, 5), (-1, 5), colors.HexColor('#DCFCE7')),
        ('FONTNAME', (0, 5), (-1, 5), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0, 5), (-1, 5), colors.HexColor('#166534')),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.5*cm))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#CBD5E1'), spaceAfter=8))
    story.append(Paragraph("Ce document est un reçu officiel SolidariCash.", center_style))

    doc.build(story)
    buffer.seek(0)
    return buffer
